"""
Add a "Woodmac Original Name" column to the Woodmac_reviewed sheet of the
Capacity Change Comparison Excel file.

The Plant column contains standardized names (output of the mapping in capacity.py).
This script does a reverse-lookup against at_lng.mapping_plant_name to recover
the original Woodmac source name for each plant.
"""

import sys
import os

import openpyxl
import pandas as pd

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from dash_shipping_lng_snd.utils.export_flow_data import engine, DB_SCHEMA

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "Capacity_Change_Comparison_20260401_162656.xlsx")
SHEET_NAME = "Woodmac_reviewed"
NEW_COL_HEADER = "Woodmac Original Name"


def fetch_plant_mapping():
    query = f"""
        SELECT country_name, source_name, plant_name
        FROM {DB_SCHEMA}.mapping_plant_name
        WHERE provider = 'woodmac' AND source_field = 'plant_name'
    """
    with engine.connect() as connection:
        mapping_df = pd.read_sql_query(query, connection)

    for col in ["country_name", "source_name", "plant_name"]:
        mapping_df[col] = mapping_df[col].fillna("").astype(str).str.strip()

    mapping_df = mapping_df[
        (mapping_df["country_name"] != "")
        & (mapping_df["source_name"] != "")
        & (mapping_df["plant_name"] != "")
    ]

    # Build reverse-lookup: (country_name.upper(), plant_name.upper()) -> source_name
    # Keep first occurrence on duplicates (same as capacity.py dedup logic)
    mapping_df["__key"] = list(
        zip(mapping_df["country_name"].str.upper(), mapping_df["plant_name"].str.upper())
    )
    mapping_df = mapping_df.drop_duplicates(subset=["__key"], keep="first")

    return dict(zip(mapping_df["__key"], mapping_df["source_name"]))


def main():
    print(f"Fetching plant mapping from database...")
    reverse_lookup = fetch_plant_mapping()
    print(f"  Loaded {len(reverse_lookup)} mapping entries.")

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    # Insert new column C (after Plant in column B), shifting existing C–E to D–F
    ws.insert_cols(3)
    ws.cell(row=1, column=3, value=NEW_COL_HEADER)

    matched = 0
    for row_idx in range(2, ws.max_row + 1):
        country = ws.cell(row=row_idx, column=1).value
        plant = ws.cell(row=row_idx, column=2).value

        if not country or not plant:
            continue

        key = (str(country).strip().upper(), str(plant).strip().upper())
        original_name = reverse_lookup.get(key)

        ws.cell(row=row_idx, column=3, value=original_name)
        if original_name:
            matched += 1

    print(f"  Matched {matched} / {ws.max_row - 1} rows to an original Woodmac name.")

    wb.save(EXCEL_PATH)
    print(f"Saved: {EXCEL_PATH}")


if __name__ == "__main__":
    main()
