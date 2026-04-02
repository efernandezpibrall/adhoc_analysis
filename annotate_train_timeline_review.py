from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


INPUT_PATH = Path("/home/efernandez/development/Github/adhoc_analysis/Train_Timeline_20260402_121207.xlsx")
OUTPUT_PATH = INPUT_PATH.with_name(f"{INPUT_PATH.stem}_reviewed.xlsx")
COMMENT_HEADER = "Codex Comments"
STATUS_GOOD = "green"
STATUS_CHECK = "yellow"
STATUS_WRONG = "red"

FILL_BY_STATUS = {
    STATUS_GOOD: PatternFill(fill_type="solid", fgColor="E2F0D9"),
    STATUS_CHECK: PatternFill(fill_type="solid", fgColor="FFF2CC"),
    STATUS_WRONG: PatternFill(fill_type="solid", fgColor="F4CCCC"),
}


def fmt_num(value):
    if value is None:
        return "n/a"
    if abs(float(value) - round(float(value))) < 1e-9:
        return str(int(round(float(value))))
    return f"{float(value):.2f}".rstrip("0").rstrip(".")


def row_value(row, name):
    return row.get(name)


def train_label(row):
    value = row_value(row, "Train")
    return "blank" if value in (None, "") else str(value)


def has_wm(row):
    return row_value(row, "Woodmac Total Capacity Added") is not None


def has_ea(row):
    return row_value(row, "Energy Aspects Total Capacity Added") is not None


def both_providers(row):
    return has_wm(row) and has_ea(row)


def aligned(a, b, tolerance=0.15):
    if a is None or b is None:
        return False
    ref = max(abs(float(a)), abs(float(b)), 1.0)
    return abs(float(a) - float(b)) / ref <= tolerance


def provider_only_phrase(row):
    if has_wm(row) and not has_ea(row):
        return "Only Woodmac shows it in the selected range, so timing confidence is lower."
    if has_ea(row) and not has_wm(row):
        return "Only Energy Aspects shows it in the selected range, so timing confidence is lower."
    return ""


def generic_comment(row, note):
    if both_providers(row):
        if aligned(row_value(row, "Woodmac Total Capacity Added"), row_value(row, "Energy Aspects Total Capacity Added")):
            return f"{note} Woodmac and Energy Aspects are close on this row; remaining differences look like normal timing/rounding."
        return f"{note} The row is directionally plausible, but the providers still disagree enough that it is worth treating the timing/capacity as medium confidence."
    return f"{note} {provider_only_phrase(row)}"


def generic_status(row):
    if both_providers(row) and aligned(
        row_value(row, "Woodmac Total Capacity Added"),
        row_value(row, "Energy Aspects Total Capacity Added"),
    ):
        return STATUS_GOOD
    return STATUS_CHECK


GENERIC_NOTES = {
    ("Australia", "Pluto"): "Public project descriptions support Pluto Train 2 at about 5 mtpa, so this scale makes sense.",
    ("Canada", "Cedar LNG"): "Cedar LNG is a single 3.3 mtpa FLNG project, so this row is structurally clean.",
    ("Canada", "Ksi Lisims LNG"): "Ksi Lisims is marketed as a 12 mtpa project, which fits two trains of roughly 6 mtpa each.",
    ("Congo", "Nguya FLNG"): "Eni describes Nguya as a 2.4 mtpa FLNG unit in Congo LNG Phase 2, so this row is well grounded publicly.",
    ("Gabon", "Gabon LNG"): "Perenco's Gabon LNG project is a 0.7 mtpa development, so this scale is sensible.",
    ("Malaysia", "PFLNG 3"): "PFLNG 3 is publicly described around 2 mtpa, so this row looks reasonable.",
    ("Mexico", "Amigo LNG"): "The two-train structure here is internally consistent and there is no obvious train-mapping red flag in the export.",
    ("Mozambique", "Coral Norte"): "Eni's Coral Norte FLNG is a single train around 3.5-3.6 mtpa, so this row is structurally sound.",
    ("Mozambique", "Mozambique LNG"): "Mozambique LNG Phase 1 is a two-train project totaling about 13.1 mtpa, so the train sizes look right.",
    ("Oman", "Marsa LNG"): "Marsa LNG is a 1 mtpa LNG bunkering project, so this row looks well aligned with public scope.",
    ("Oman", "Oman LNG"): "A 3.8 mtpa increment is plausible for the Oman LNG/Qalhat complex's added line, so the row is reasonable.",
    ("Qatar", "North Field East"): "QatarEnergy's North Field East is four mega-trains totaling 32 mtpa, so these row sizes are exactly the right order of magnitude.",
    ("Qatar", "North Field South"): "QatarEnergy's North Field South is two mega-trains totaling 16 mtpa, so these row sizes are exactly the right order of magnitude.",
    ("United Arab Emirates", "Ruwais LNG"): "ADNOC's Ruwais LNG is two trains totaling 9.6 mtpa, so 4.8 per train is the expected public structure.",
    ("United States", "Golden Pass"): "Golden Pass is a three-train project of about 18 mtpa total, so these row sizes are well grounded publicly.",
    ("United States", "Port Arthur LNG"): "Port Arthur Phases 1 and 2 are four trains totaling about 26 mtpa, so these row sizes make sense.",
    ("United States", "Rio Grande LNG"): "Rio Grande Trains 1-5 are publicly marketed around 5.4-5.9 mtpa each, so the structure here is sensible.",
}


def plant_comment(row, group):
    country = row_value(row, "Country")
    plant = row_value(row, "Plant")
    train = train_label(row)
    key = (country, plant)

    if key in GENERIC_NOTES:
        return generic_comment(row, GENERIC_NOTES[key])

    if key == ("Algeria", "Algeria LNG"):
        return "Three identical 0.37 mtpa rows look more like modeled brownfield increments than clearly public new trains. Only Woodmac carries them, so confidence on the exact train split is low."

    if key == ("Argentina", "Argentina LNG"):
        if train in {"1", "2"}:
            return "These first two rows make sense. Public YPF material supports the first two FLNG units at roughly 12 mtpa total, so about 6 mtpa each is plausible. The gap is that Energy Aspects also carries a later third unit while Woodmac does not yet."
        return "This looks like Energy Aspects' later third FLNG unit / expansion. A ~5 mtpa row is directionally plausible, but Woodmac does not yet show it, so the plant total diverges for completeness rather than obvious bad sizing."

    if key == ("Argentina", "Golfo San Matias FLNG"):
        return "Providers match very closely on this row, so there is no obvious train-mapping problem in the export. The plant structure is unusual, but the row itself looks internally consistent."

    if key == ("Canada", "LNG Canada"):
        if train == "2":
            return "Reasonable. LNG Canada Phase 1 is two trains totaling 14 mtpa, so 7 mtpa for Train 2 fits. Energy Aspects is probably missing here because Train 2's positive addition likely fell outside the exported date window."
        return "This looks like a speculative Phase 2 expansion train. A 7 mtpa block is plausible, but public Phase 2 is still not sanctioned, so certainty is lower even though both providers carry it."

    if key == ("Canada", "Port Edward"):
        return "The 0.3 mtpa size is plausible for the small-scale Port Edward concept, but recent public trackers treat the project as inactive/cancelled. This looks more like a stale forward project than a high-confidence live development."

    if key == ("Canada", "Quebec LNG"):
        return "Two trains around 5.65 mtpa each fit the old Energie Saguenay 10.5-11 mtpa concept, but the project was federally rejected on February 7, 2022 and appears cancelled/stale. Capacity structure is plausible; timeline confidence is not."

    if key == ("Canada", "Tilbury"):
        return "The plant total here is broadly plausible for FortisBC's Tilbury Phase 2 liquefaction expansion, but public material does not clearly market two discrete 1.4 mtpa trains. Treat the two-train split as model-driven rather than publicly confirmed."

    if key == ("Canada", "Woodfibre LNG"):
        if train == "1":
            return "This export is stale. Publicly Woodfibre is a single-train 2.1 mtpa facility, and the live mapping has already been corrected so Energy Aspects' 1.05 + 1.05 steps roll up to Train 1. In this older export, EA still shows only the first half-step on Train 1."
        return "This separate Train 2 does not make physical sense as a distinct marketed train. It is the second 1.05 mtpa step of the same single-train 2.1 mtpa Woodfibre project, and the live mapping has since been corrected to map it back to Train 1."

    if key == ("Congo", "Tango FLNG"):
        return "This row does not make sense as shown. Eni's public Congo LNG structure is Tango FLNG at 0.6 mtpa and Nguya FLNG at 2.4 mtpa; a blank-train Tango row at 2.4 mtpa is likely an alias/modeling error, not a real Tango train."

    if key == ("Djibouti", "Ethiopia-Djibouti LNG"):
        return "A 3 mtpa single-train first phase is directionally plausible for this long-discussed project, but public execution remains weak and only Energy Aspects carries it. Treat it as low-confidence forward modeling."

    if key == ("Indonesia", "Abadi"):
        if train == "1":
            return "Public current Abadi concept is about 9.5 mtpa total, which fits two trains near 4.75 mtpa. This row is directionally fine, but Woodmac is slightly smaller and seems to carry only the first train."
        return "This second 4.75 mtpa row is consistent with the current two-train ~9.5 mtpa Abadi concept. The mismatch is that Woodmac does not yet carry this second train in the export."

    if key == ("Indonesia", "Arun"):
        return "Six trains around 2.05 mtpa match the historical six-train Arun configuration. As a future-addition row set this looks like a restart/revival scenario modeled only in Woodmac, so confidence is lower."

    if key == ("Indonesia", "Genting FLNG"):
        if train == "1":
            return "The overall 1.2 mtpa project size is plausible. Energy Aspects keeps the full 1.2 on Train 1, while Woodmac splits the same facility into two 0.6 units, so this is mainly a provider granularity issue."
        return "This Woodmac-only 0.6 row is likely the second half of the same 1.2 mtpa FLNG facility rather than a clearly distinct marketed train. Public descriptions emphasize one 1.2 mtpa FLNG project."

    if key == ("Indonesia", "Sengkang LNG"):
        return "A 0.5 mtpa single-train concept is directionally plausible, but only Woodmac shows it and public train-level detail is limited. Treat it as low-confidence forward modeling."

    if key == ("Libya", "Marsa El Brega"):
        return "A two-row 2.6 mtpa total may be plausible for a Libyan restart/expansion concept, but public train-level confirmation is limited and only Woodmac carries it. Confidence on the exact split between Train 1 and Train 2 is low."

    if key == ("Mauritania", "Tortue FLNG"):
        return "A roughly 2.4-2.5 mtpa second liquefaction unit is directionally consistent with Greater Tortue Ahmeyim Phase 2 concepts. The issue is maturity: public Phase 2 remains under evaluation, so this is plausible but low-confidence."

    if key == ("Mexico", "Costa Azul LNG"):
        if train == "1":
            return "Looks good. Public Sempra material clearly defines ECA LNG Phase 1 as a single 3.25 mtpa train, and both providers match this row closely."
        return "These Woodmac-only 5.5 mtpa rows look like the two planned ECA Phase 2 trains. That expansion concept is plausible publicly, but Energy Aspects does not yet carry it in this export, so treat it as forward/speculative coverage rather than a mapping problem."

    if key == ("Mexico", "Saguaro Energía"):
        if train == "1":
            return "This row is directionally correct. Mexico Pacific currently describes Saguaro as three trains totaling 15 mtpa, with the first two around 4.7 mtpa each. Woodmac's 5.0 vs EA's 4.7 is close enough."
        if train == "2":
            return "Public structure supports a second Saguaro train of roughly this size, but Energy Aspects has not carried it yet in this export. This looks like provider completeness rather than bad train mapping."
        return "Public structure also supports a third Saguaro train, but it is still later-phase capacity and Energy Aspects has not carried it here. Plausible, but still speculative."

    if key == ("Mexico", "Vista Pacifico LNG"):
        return "A single 4 mtpa train is plausible for the Vista Pacifico concept, but only Woodmac carries it in this export. Cross-provider confidence is lower."

    if key == ("Mozambique", "Rovuma"):
        return "The overall scale is plausible, but the train structure is not. Public Rovuma LNG is generally described as two trains with combined capacity a little above 15 mtpa, whereas this export shows twelve 1.5 mtpa rows. Treat these train numbers as provider model blocks, not real marketed trains."

    if key == ("Nigeria", "Ace FLNG"):
        return "Reasonable in size. Public project disclosures around Ace FLNG point to a 3 mtpa FLNG concept, so the capacity itself makes sense. Only Woodmac carries it."

    if key == ("Nigeria", "NLNG"):
        if train == "7":
            return "Public NLNG messaging treats Train 7 as one expansion project that lifts capacity from 22 to 30 mtpa. Energy Aspects' ~7.6 row and Woodmac's 4.2 row together suggest Woodmac separates the liquefaction train from part of the debottlenecking gain, so the exact split is provider-model specific."
        return "This Woodmac-only 3.5 row does not look like a publicly marketed standalone Train 8. It is more likely Woodmac's way of parking the debottlenecking benefit that public sources usually bundle into Train 7."

    if key == ("Nigeria", "Nigeria FLNG"):
        return "A 2.1 mtpa single-train FLNG is directionally plausible for a small floating concept, but public project definitions remain fluid and only Energy Aspects carries it. Medium-to-low confidence."

    if key == ("Papua New Guinea", "Papua LNG"):
        if train == "4":
            return "Reasonable in structure. Papua LNG's current concept uses four 1 mtpa e-trains; the only issue is that Woodmac does not yet show this fourth train in the export."
        return "Reasonable. Current TotalEnergies concept uses four e-trains totaling 4 mtpa, so 1 mtpa per train makes sense. Any provider difference here is mostly timing."

    if key == ("Qatar", "Ras Laffan"):
        return "This row does not carry any first effective date or capacity added from either provider. It looks like a table artifact rather than a useful selected-range timeline row and could probably be filtered out."

    if key == ("Russia", "Arctic LNG 2"):
        return "Reasonable. Public Arctic LNG 2 design is three 6.6 mtpa trains, so a Train 3 row at 6.6 is structurally correct. The lack of Woodmac support is more about provider coverage than train logic."

    if key == ("Russia", "Far East LNG (Sakhalin 1)"):
        return "A single train around 6.2 mtpa is directionally plausible for the Far East LNG concept, but only Energy Aspects carries it and project certainty remains limited."

    if key == ("Russia", "Obsky"):
        return "Reasonable. Public Obsky LNG concepts have been around 6 mtpa total in two trains, so 3 + 3 fits the structure. Only Energy Aspects carries it, so timing remains lower confidence."

    if key == ("Senegal", "Yakaar-Teranga LNG"):
        return "A roughly 2.5 mtpa FLNG export unit is directionally plausible, but current public signals are weak: Kosmos has discussed an offshore LNG concept while project ownership/partnering remains unsettled. Treat it as plausible but low-confidence."

    if key == ("Suriname", "Block 52 LNG"):
        return "This row looks too early/speculative. Public Block 52 information still centers on upstream discoveries and partner changes rather than a defined LNG train, so a 3 mtpa Train 1 should be treated as very low-confidence forward modeling."

    if key == ("Tanzania", "Tanzania LNG"):
        return "The size is plausible as a phase or single-train simplification inside a much larger Tanzania LNG concept, but public train-level definition is still vague and only Woodmac carries it."

    if key == ("United States", "CP2 LNG Phase 1"):
        return "These are not conventional large trains. Venture Global projects use many modular liquefaction blocks; Woodmac models ~0.78 mtpa sub-trains while Energy Aspects aggregates roughly pairs into ~1.54 mtpa blocks. Treat the numbering here as provider-specific blocks, not public market-facing trains."

    if key == ("United States", "CP2 LNG Phase 2"):
        return "Same issue as CP2 Phase 1: these are provider model blocks, not conventional marketed trains. The ~0.78 mtpa Woodmac rows are plausible as Venture Global modules, but train numbering here should not be read literally."

    if key == ("United States", "Cameron LNG"):
        return "This looks like Cameron Phase 2 / a first expansion train. A ~6.75 mtpa row is plausible publicly, but only Woodmac carries it here."

    if key == ("United States", "Commonwealth"):
        if train in {"1", "2"}:
            return "This is plausible. Commonwealth is publicly structured as six liquefaction facilities/trains totaling roughly 8.4-9.5 mtpa, so about 1.58 mtpa each makes sense. Energy Aspects currently carries only the first two rows, which makes the plant look incomplete rather than wrongly mapped."
        return "This row is plausible under Commonwealth's six-train modular structure. The real issue is provider completeness: Woodmac carries all six trains while Energy Aspects only shows the first two in this export."

    if key == ("United States", "Corpus Christi"):
        return "These tiny 0.11 mtpa rows look like existing-train debottleneck increments at the base Corpus Christi terminal, not new full trains. Structurally that is plausible, but only Woodmac carries it."

    if key == ("United States", "Corpus Christi Stage 3"):
        if train in {"5", "6", "7"}:
            return "Reasonable. Cheniere's Stage 3 is seven midscale trains, and this row size around 1.43-1.49 mtpa per train is plausible. Woodmac and Energy Aspects are broadly aligned here."
        if train in {"8", "9"}:
            return "Reasonable as later midscale trains. Slight Woodmac/EA spread is acceptable; public Cheniere updates support additional midscale capacity around this scale."
        return "This pseudo-train is best read as debottlenecking / non-train incremental capacity, not a physical train. That makes sense because Cheniere publicly bundles Stage 3 with extra debottlenecking capacity."

    if key == ("United States", "Delfin FLNG"):
        if train == "1":
            return "Reasonable. Public Delfin materials support FLNG vessels of about 4.4 mtpa each, so this row makes sense and both providers align on the first vessel."
        return "This Woodmac-only second 4.4 mtpa row is still plausible because Delfin is publicly marketed as a multi-vessel project. The issue is that Energy Aspects has not yet carried the second vessel in this export."

    if key == ("United States", "Freeport"):
        return "This looks like the proposed fourth train. A ~4.6 mtpa row is plausible publicly, but only Woodmac carries it here."

    if key == ("United States", "Jacksonville LNG"):
        return "These 0.33 mtpa rows look like small modular units rather than conventional large export trains. Plausible as a small-scale expansion concept, but only Woodmac carries them."

    if key == ("United States", "Lake Charles Export"):
        return "Three trains around 5.48 mtpa fit the long-standing Lake Charles export concept. Only Woodmac shows it here, so the main uncertainty is timing/advancement, not the capacity split."

    if key == ("United States", "Louisiana LNG"):
        return "These rows probably refer to the same asset as 'Woodside Louisiana LNG' under the project's newer branding. The five 5.5 mtpa trains are plausible for the full 27.6 mtpa permitted facility, but plant mapping is not harmonized across providers, so the comparison is artificially split."

    if key == ("United States", "Magnolia LNG"):
        return "Four trains around 2.2 mtpa align with Magnolia's long-standing 8.8 mtpa four-train concept. Only Woodmac carries them, so confidence is lower on timing rather than structure."

    if key == ("United States", "Plaquemines"):
        return "Treat these as provider model blocks, not conventional marketed trains. Venture Global uses many modular trains/blocks at Plaquemines, so 1.6-1.66 mtpa row sizes are plausible, but Woodmac absence means there is no cross-provider confirmation in this export."

    if key == ("United States", "Sabine Pass"):
        if train in {"1", "2", "3", "4", "5", "6"}:
            return "These are small debottleneck increments on the existing six-train plant, not new trains. That is structurally plausible, but only Woodmac carries the increment in this export."
        if train in {"7", "8", "9"}:
            return "These look like Stage 5 expansion trains. Public Cheniere material supports up to three new large trains around 6-6.5 mtpa, so the row is plausible."
        return "This pseudo-train is likely re-liquefaction / debottleneck capacity, not a physical train, which matches Cheniere's public expansion description."

    if key == ("United States", "Texas LNG"):
        return "Two trains at 2 mtpa each are plausible for Texas LNG's 4 mtpa concept. Only Woodmac carries them here."

    if key == ("United States", "Woodside Louisiana LNG"):
        return "These rows also look like the Louisiana LNG / former Driftwood asset, but under Woodside's current name. The three 5.5 mtpa trains match the sanctioned 16.5 mtpa phase 1; the main issue is that Energy Aspects still appears under the separate plant name 'Louisiana LNG'."

    if has_wm(row) and has_ea(row):
        return "The row is internally consistent because both providers carry it. Public train-level detail is thinner here, but nothing in the export itself suggests a mapping issue."

    return f"The size and train assignment are directionally plausible, but public train-level detail is limited and {provider_only_phrase(row).lower()}"


def plant_status(row, group):
    country = row_value(row, "Country")
    plant = row_value(row, "Plant")
    train = train_label(row)
    key = (country, plant)

    if key in GENERIC_NOTES:
        return generic_status(row)

    if key == ("Canada", "Woodfibre LNG"):
        return STATUS_WRONG

    if key == ("Congo", "Tango FLNG"):
        return STATUS_WRONG

    if key == ("Qatar", "Ras Laffan"):
        return STATUS_WRONG

    if key in {
        ("United States", "Louisiana LNG"),
        ("United States", "Woodside Louisiana LNG"),
    }:
        return STATUS_WRONG

    if key == ("Mexico", "Costa Azul LNG"):
        return STATUS_GOOD if train == "1" else STATUS_CHECK

    if key == ("Canada", "LNG Canada"):
        return STATUS_CHECK

    if key == ("Argentina", "Argentina LNG"):
        return STATUS_CHECK

    if key == ("Canada", "Port Edward"):
        return STATUS_CHECK

    if key == ("Canada", "Quebec LNG"):
        return STATUS_CHECK

    if key == ("Canada", "Tilbury"):
        return STATUS_CHECK

    if key == ("Djibouti", "Ethiopia-Djibouti LNG"):
        return STATUS_CHECK

    if key == ("Indonesia", "Abadi"):
        return STATUS_CHECK

    if key == ("Indonesia", "Arun"):
        return STATUS_CHECK

    if key == ("Indonesia", "Genting FLNG"):
        return STATUS_CHECK

    if key == ("Indonesia", "Sengkang LNG"):
        return STATUS_CHECK

    if key == ("Libya", "Marsa El Brega"):
        return STATUS_CHECK

    if key == ("Mauritania", "Tortue FLNG"):
        return STATUS_CHECK

    if key == ("Mexico", "Saguaro Energía"):
        return STATUS_CHECK

    if key == ("Mexico", "Vista Pacifico LNG"):
        return STATUS_CHECK

    if key == ("Mozambique", "Rovuma"):
        return STATUS_CHECK

    if key == ("Nigeria", "Ace FLNG"):
        return STATUS_CHECK

    if key == ("Nigeria", "NLNG"):
        return STATUS_CHECK

    if key == ("Nigeria", "Nigeria FLNG"):
        return STATUS_CHECK

    if key == ("Papua New Guinea", "Papua LNG"):
        return STATUS_CHECK if train == "4" else STATUS_GOOD

    if key == ("Russia", "Arctic LNG 2"):
        return STATUS_CHECK

    if key == ("Russia", "Far East LNG (Sakhalin 1)"):
        return STATUS_CHECK

    if key == ("Russia", "Obsky"):
        return STATUS_CHECK

    if key == ("Senegal", "Yakaar-Teranga LNG"):
        return STATUS_CHECK

    if key == ("Suriname", "Block 52 LNG"):
        return STATUS_CHECK

    if key == ("Tanzania", "Tanzania LNG"):
        return STATUS_CHECK

    if key == ("United States", "CP2 LNG Phase 1"):
        return STATUS_CHECK

    if key == ("United States", "CP2 LNG Phase 2"):
        return STATUS_CHECK

    if key == ("United States", "Cameron LNG"):
        return STATUS_CHECK

    if key == ("United States", "Commonwealth"):
        return STATUS_CHECK

    if key == ("United States", "Corpus Christi"):
        return STATUS_CHECK

    if key == ("United States", "Corpus Christi Stage 3"):
        return STATUS_CHECK if train == "999" else STATUS_GOOD

    if key == ("United States", "Delfin FLNG"):
        return STATUS_GOOD if train == "1" else STATUS_CHECK

    if key == ("United States", "Freeport"):
        return STATUS_CHECK

    if key == ("United States", "Jacksonville LNG"):
        return STATUS_CHECK

    if key == ("United States", "Lake Charles Export"):
        return STATUS_CHECK

    if key == ("United States", "Magnolia LNG"):
        return STATUS_CHECK

    if key == ("United States", "Plaquemines"):
        return STATUS_CHECK

    if key == ("United States", "Sabine Pass"):
        return STATUS_CHECK

    if key == ("United States", "Texas LNG"):
        return STATUS_CHECK

    return generic_status(row)


def main():
    wb = load_workbook(INPUT_PATH)
    ws = wb[wb.sheetnames[0]]
    headers = [cell.value for cell in ws[1]]
    header_index = {name: idx for idx, name in enumerate(headers)}

    groups = {}
    row_dicts = []
    for excel_row in range(2, ws.max_row + 1):
        values = [ws.cell(row=excel_row, column=col).value for col in range(1, ws.max_column + 1)]
        row = dict(zip(headers, values))
        row_dicts.append((excel_row, row))
        groups.setdefault((row["Country"], row["Plant"]), []).append(row)

    comment_col = ws.max_column + 1
    ws.cell(row=1, column=comment_col).value = COMMENT_HEADER
    ws.cell(row=1, column=comment_col).font = Font(bold=True)
    ws.cell(row=1, column=comment_col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[ws.cell(row=1, column=comment_col).column_letter].width = 105

    for excel_row, row in row_dicts:
        comment = plant_comment(row, groups[(row["Country"], row["Plant"])])
        status = plant_status(row, groups[(row["Country"], row["Plant"])])
        cell = ws.cell(row=excel_row, column=comment_col)
        cell.value = comment
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        row_fill = FILL_BY_STATUS[status]
        for col in range(1, comment_col + 1):
            ws.cell(row=excel_row, column=col).fill = row_fill

    wb.save(OUTPUT_PATH)
    print(f"Reviewed workbook written to: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
