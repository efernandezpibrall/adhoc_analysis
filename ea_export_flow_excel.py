#!/usr/bin/env python3
"""
Generate a monthly Energy Aspects LNG export Excel file by origin country.

Workbook sheets:
- Exports Flow

Output columns:
- Month
- Total MMTPA
- US
- Qatar
- United Arab Emirates
- Australia
- Russia
- Canada
- Mozambique
- Rest of the World
"""

from __future__ import annotations

import argparse
import configparser
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine


CONFIG_PATH = Path("/home/efernandez/development/Github/config.ini")
OUTPUT_DIR = Path("/home/efernandez/development/Github/adhoc_analysis")
SHEET_NAME = "Exports Flow"

COUNTRY_LABELS = {
    "US": "US",
    "United States": "US",
    "Qatar": "Qatar",
    "UAE": "United Arab Emirates",
    "United Arab Emirates": "United Arab Emirates",
    "Australia": "Australia",
    "Russia": "Russia",
    "Canada": "Canada",
    "Mozambique": "Mozambique",
}

OUTPUT_COLUMNS = [
    "Month",
    "Total MMTPA",
    "US",
    "Qatar",
    "United Arab Emirates",
    "Australia",
    "Russia",
    "Canada",
    "Mozambique",
    "Rest of the World",
]

EA_EXPORT_QUERY = """
WITH latest_snapshot AS (
    SELECT MAX(upload_timestamp_utc) AS upload_timestamp_utc
    FROM at_lng.ea_values
),
export_mappings AS (
    SELECT
        CAST(dataset_id AS TEXT) AS dataset_id,
        country,
        unit,
        frequency
    FROM at_lng.fundamentals_ea_lng_balance_datasets
    WHERE aspect = 'exports'
      AND category_subtype = 'LNG'
      AND frequency = 'monthly'
      AND unit = 'Mt'
      AND country IS NOT NULL
      AND country <> ''
),
latest_values AS (
    SELECT
        a.dataset_id,
        a.date::date AS month,
        a.value
    FROM at_lng.ea_values a
    JOIN latest_snapshot s
        ON a.upload_timestamp_utc = s.upload_timestamp_utc
    WHERE a.dataset_id IN (SELECT dataset_id FROM export_mappings)
)
SELECT
    v.month,
    m.country AS country_name,
    SUM(v.value * 12.0) AS total_mmtpa
FROM latest_values v
JOIN export_mappings m
    ON v.dataset_id = m.dataset_id
GROUP BY v.month, m.country
ORDER BY v.month, m.country
"""


def parse_args() -> argparse.Namespace:
    timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
    default_output = OUTPUT_DIR / f"ea_export_flow_monthly_{timestamp}.xlsx"

    parser = argparse.ArgumentParser(
        description="Generate a monthly Energy Aspects LNG export Excel workbook."
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=default_output,
        help="Excel output path. Defaults to adhoc_analysis/ea_export_flow_monthly_<timestamp>.xlsx",
    )
    return parser.parse_args()


def get_engine():
    config = configparser.ConfigParser(interpolation=None)
    config.read(CONFIG_PATH)
    return create_engine(config["DATABASE"]["CONNECTION_STRING"], pool_pre_ping=True)


def fetch_export_data(conn) -> pd.DataFrame:
    return pd.read_sql_query(EA_EXPORT_QUERY, conn)


def build_country_matrix(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        raise ValueError("The Energy Aspects export query returned no data.")

    report_df = df.copy()
    report_df["month"] = pd.to_datetime(report_df["month"])
    report_df["country_bucket"] = report_df["country_name"].map(COUNTRY_LABELS).fillna(
        "Rest of the World"
    )

    bucketed = (
        report_df.groupby(["month", "country_bucket"], as_index=False)["total_mmtpa"]
        .sum()
        .sort_values(["month", "country_bucket"])
    )

    pivot = (
        bucketed.pivot(index="month", columns="country_bucket", values="total_mmtpa")
        .fillna(0.0)
        .sort_index()
    )

    month_index = pd.date_range(
        start=report_df["month"].min(),
        end=report_df["month"].max(),
        freq="MS",
    )
    pivot = pivot.reindex(month_index, fill_value=0.0)
    pivot.index.name = "month"

    for column in OUTPUT_COLUMNS[2:]:
        if column not in pivot.columns:
            pivot[column] = 0.0

    pivot["Total MMTPA"] = pivot[OUTPUT_COLUMNS[2:]].sum(axis=1)

    result = pivot.reset_index().rename(columns={"month": "Month"})
    result = result[OUTPUT_COLUMNS]

    numeric_columns = [column for column in OUTPUT_COLUMNS if column != "Month"]
    result[numeric_columns] = result[numeric_columns].round(2)

    return result


def format_worksheet(worksheet) -> None:
    header_font = Font(bold=True)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet[1]:
        cell.font = header_font

    for cell in worksheet["A"][1:]:
        cell.number_format = "yyyy-mm"

    for row in worksheet.iter_rows(min_row=2, min_col=2):
        for cell in row:
            cell.number_format = "0.00"

    for column_cells in worksheet.columns:
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        worksheet.column_dimensions[
            get_column_letter(column_cells[0].column)
        ].width = max_length + 2


def export_to_excel(df: pd.DataFrame, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(
        output_path,
        engine="openpyxl",
        datetime_format="yyyy-mm",
        date_format="yyyy-mm",
    ) as writer:
        df.to_excel(writer, sheet_name=SHEET_NAME, index=False)
        format_worksheet(writer.sheets[SHEET_NAME])

    return output_path


def main() -> None:
    args = parse_args()
    engine = get_engine()

    try:
        with engine.connect() as conn:
            print("Fetching Energy Aspects LNG export flow data...")
            raw_df = fetch_export_data(conn)
    finally:
        engine.dispose()

    print(
        f"Retrieved {len(raw_df):,} country-month rows from "
        f"{raw_df['month'].min()} to {raw_df['month'].max()}."
    )

    print("Building Exports Flow matrix...")
    report_df = build_country_matrix(raw_df)

    print(f"Writing Excel file to {args.output}...")
    output_path = export_to_excel(report_df, args.output)

    print("Done.")
    print(f"Output file: {output_path}")
    print(
        f"{SHEET_NAME}: {len(report_df):,} rows | "
        f"Months: {report_df['Month'].min():%Y-%m} to {report_df['Month'].max():%Y-%m}"
    )


if __name__ == "__main__":
    main()
