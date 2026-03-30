#!/usr/bin/env python3
"""
Generate a monthly WoodMac LNG export Excel file by origin country.

Workbook sheets:
- Exports Flow
- Exports Capacity
- Exports Maintenance

Output columns:
- Month
- Total MMTPA
- US
- Qatar
- United Arab Emirates
- Australia
- Russia
- Canada
- Mozambique
- Rest of the World
"""

from __future__ import annotations

import argparse
import configparser
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine


CONFIG_PATH = Path("/home/efernandez/development/Github/config.ini")
OUTPUT_DIR = Path("/home/efernandez/development/Github/adhoc_analysis")
SOURCE_TABLE = "at_lng.woodmac_gas_imports_exports_monthly__mmtpa"

COUNTRY_LABELS = {
    "United States": "US",
    "Qatar": "Qatar",
    "United Arab Emirates": "United Arab Emirates",
    "Australia": "Australia",
    "Russia": "Russia",
    "Canada": "Canada",
    "Mozambique": "Mozambique",
}

OUTPUT_COLUMNS = [
    "Month",
    "Total MMTPA",
    "US",
    "Qatar",
    "United Arab Emirates",
    "Australia",
    "Russia",
    "Canada",
    "Mozambique",
    "Rest of the World",
]

METRICS = {
    "Flow": "Exports Flow",
}

CAPACITY_SHEET_NAME = "Exports Capacity"
MAINTENANCE_SHEET_NAME = "Exports Maintenance"

MONTH_YEAR_REGEX = (
    "(January|February|March|April|May|June|July|August|September|October|"
    "November|December)\\s+(\\d{4})"
)

QUERY_TEMPLATE = """
WITH latest_short_term_market AS (
    SELECT market_outlook
    FROM {source_table}
    WHERE release_type = 'Short Term Outlook'
      AND direction = 'Export'
      AND measured_at = 'Exit'
      AND metric_name = '{metric_name}'
    GROUP BY market_outlook
    ORDER BY TO_DATE(
        (regexp_match(market_outlook, '{month_year_regex}'))[1]
        || ' ' ||
        (regexp_match(market_outlook, '{month_year_regex}'))[2],
        'Month YYYY'
    ) DESC NULLS LAST,
    MAX(publication_date::timestamp) DESC
    LIMIT 1
),
latest_long_term_market AS (
    SELECT market_outlook
    FROM {source_table}
    WHERE release_type = 'Long Term Outlook'
      AND direction = 'Export'
      AND measured_at = 'Exit'
      AND metric_name = '{metric_name}'
    GROUP BY market_outlook
    ORDER BY TO_DATE(
        (regexp_match(market_outlook, '{month_year_regex}'))[1]
        || ' ' ||
        (regexp_match(market_outlook, '{month_year_regex}'))[2],
        'Month YYYY'
    ) DESC NULLS LAST,
    MAX(publication_date::timestamp) DESC
    LIMIT 1
),
short_term AS (
    SELECT
        start_date::date AS month,
        country_name,
        SUM(metric_value) AS total_mmtpa
    FROM {source_table}
    WHERE market_outlook = (SELECT market_outlook FROM latest_short_term_market)
      AND release_type = 'Short Term Outlook'
      AND direction = 'Export'
      AND measured_at = 'Exit'
      AND metric_name = '{metric_name}'
    GROUP BY start_date::date, country_name
    HAVING SUM(metric_value) > 0
),
short_term_max_month AS (
    SELECT MAX(month) AS max_month
    FROM short_term
),
long_term_raw AS (
    SELECT
        start_date::date AS month,
        country_name,
        SUM(metric_value) AS total_mmtpa
    FROM {source_table}
    WHERE market_outlook = (SELECT market_outlook FROM latest_long_term_market)
      AND release_type = 'Long Term Outlook'
      AND direction = 'Export'
      AND measured_at = 'Exit'
      AND metric_name = '{metric_name}'
    GROUP BY start_date::date, country_name
    HAVING SUM(metric_value) > 0
),
long_term AS (
    SELECT month, country_name, total_mmtpa
    FROM long_term_raw
    WHERE month > COALESCE(
        (SELECT max_month FROM short_term_max_month),
        DATE '1900-01-01'
    )
)
SELECT month, country_name, total_mmtpa
FROM short_term
UNION ALL
SELECT month, country_name, total_mmtpa
FROM long_term
ORDER BY month, country_name
"""

MAINTENANCE_QUERY = """
WITH combined_maintenance AS (
    SELECT
        plant_name,
        country_name,
        lng_train_name_short,
        year,
        month,
        year_actual_forecast,
        SUM(metric_value) AS total_mmtpa
    FROM (
        SELECT
            plant_name,
            country_name,
            lng_train_name_short,
            year,
            month,
            year_actual_forecast,
            metric_value
        FROM at_lng.woodmac_lng_plant_train_monthly_unplanned_downtime_mta
        WHERE metric_value > 0

        UNION ALL

        SELECT
            plant_name,
            country_name,
            lng_train_name_short,
            year,
            month,
            year_actual_forecast,
            metric_value
        FROM at_lng.woodmac_lng_plant_train_monthly_planned_maintenance_mta
        WHERE metric_value > 0
    ) maintenance_data
    GROUP BY
        plant_name,
        country_name,
        lng_train_name_short,
        year,
        month,
        year_actual_forecast
),
country_monthly AS (
    SELECT
        MAKE_DATE(year::int, month::int, 1) AS month,
        country_name,
        SUM(total_mmtpa) AS total_mmtpa
    FROM combined_maintenance
    GROUP BY MAKE_DATE(year::int, month::int, 1), country_name
)
SELECT month, country_name, total_mmtpa
FROM country_monthly
ORDER BY month, country_name
"""

CAPACITY_QUERY = """
WITH latest_plant_summary AS (
    SELECT DISTINCT ON (id_plant)
        id_plant,
        country_name
    FROM at_lng.woodmac_lng_plant_summary
    ORDER BY id_plant, upload_timestamp_utc DESC
),
latest_monthly_capacity AS (
    SELECT DISTINCT ON (id_plant, id_lng_train, year, month)
        id_plant,
        id_lng_train,
        year,
        month,
        metric_value
    FROM at_lng.woodmac_lng_plant_monthly_capacity_nominal_mta
    ORDER BY id_plant, id_lng_train, year, month, upload_timestamp_utc DESC
),
country_monthly AS (
    SELECT
        MAKE_DATE(c.year::int, c.month::int, 1) AS month,
        p.country_name,
        SUM(c.metric_value) AS total_mmtpa
    FROM latest_monthly_capacity c
    JOIN latest_plant_summary p
        ON c.id_plant = p.id_plant
    GROUP BY MAKE_DATE(c.year::int, c.month::int, 1), p.country_name
)
SELECT month, country_name, total_mmtpa
FROM country_monthly
ORDER BY month, country_name
"""


def parse_args() -> argparse.Namespace:
    timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
    default_output = OUTPUT_DIR / f"woodmac_export_flow_monthly_{timestamp}.xlsx"

    parser = argparse.ArgumentParser(
        description="Generate a monthly WoodMac LNG export Excel workbook."
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=default_output,
        help="Excel output path. Defaults to adhoc_analysis/woodmac_export_flow_monthly_<timestamp>.xlsx",
    )
    return parser.parse_args()


def build_query(metric_name: str) -> str:
    return QUERY_TEMPLATE.format(
        source_table=SOURCE_TABLE,
        metric_name=metric_name,
        month_year_regex=MONTH_YEAR_REGEX,
    )


def get_engine():
    config = configparser.ConfigParser(interpolation=None)
    config.read(CONFIG_PATH)
    return create_engine(config["DATABASE"]["CONNECTION_STRING"], pool_pre_ping=True)


def fetch_metric_data(conn, metric_name: str) -> pd.DataFrame:
    return pd.read_sql_query(build_query(metric_name), conn)


def fetch_maintenance_data(conn) -> pd.DataFrame:
    return pd.read_sql_query(MAINTENANCE_QUERY, conn)


def fetch_capacity_data(conn) -> pd.DataFrame:
    return pd.read_sql_query(CAPACITY_QUERY, conn)


def build_country_matrix(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        raise ValueError("The WoodMac query returned no data.")

    report_df = df.copy()
    report_df["month"] = pd.to_datetime(report_df["month"])
    report_df["country_bucket"] = report_df["country_name"].map(COUNTRY_LABELS).fillna(
        "Rest of the World"
    )

    bucketed = (
        report_df.groupby(["month", "country_bucket"], as_index=False)["total_mmtpa"]
        .sum()
        .sort_values(["month", "country_bucket"])
    )

    pivot = (
        bucketed.pivot(index="month", columns="country_bucket", values="total_mmtpa")
        .fillna(0.0)
        .sort_index()
    )

    month_index = pd.date_range(
        start=report_df["month"].min(),
        end=report_df["month"].max(),
        freq="MS",
    )
    pivot = pivot.reindex(month_index, fill_value=0.0)
    pivot.index.name = "month"

    for column in OUTPUT_COLUMNS[2:]:
        if column not in pivot.columns:
            pivot[column] = 0.0

    pivot["Total MMTPA"] = pivot[OUTPUT_COLUMNS[2:]].sum(axis=1)

    result = pivot.reset_index().rename(columns={"month": "Month"})
    result = result[OUTPUT_COLUMNS]

    numeric_columns = [column for column in OUTPUT_COLUMNS if column != "Month"]
    result[numeric_columns] = result[numeric_columns].round(2)

    return result


def format_worksheet(worksheet) -> None:
    header_font = Font(bold=True)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet[1]:
        cell.font = header_font

    for cell in worksheet["A"][1:]:
        cell.number_format = "yyyy-mm"

    for row in worksheet.iter_rows(min_row=2, min_col=2):
        for cell in row:
            cell.number_format = "0.00"

    for column_cells in worksheet.columns:
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        worksheet.column_dimensions[
            get_column_letter(column_cells[0].column)
        ].width = max_length + 2


def export_to_excel(sheet_data: dict[str, pd.DataFrame], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(
        output_path,
        engine="openpyxl",
        datetime_format="yyyy-mm",
        date_format="yyyy-mm",
    ) as writer:
        for sheet_name, df in sheet_data.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            format_worksheet(writer.sheets[sheet_name])

    return output_path


def main() -> None:
    args = parse_args()

    sheet_data: dict[str, pd.DataFrame] = {}
    engine = get_engine()

    try:
        with engine.connect() as conn:
            for metric_name, sheet_name in METRICS.items():
                print(f"Fetching WoodMac LNG export {metric_name.lower()} data...")
                raw_df = fetch_metric_data(conn, metric_name)
                print(
                    f"Retrieved {len(raw_df):,} country-month rows from "
                    f"{raw_df['month'].min()} to {raw_df['month'].max()}."
                )

                print(f"Building {sheet_name} matrix...")
                sheet_data[sheet_name] = build_country_matrix(raw_df)

            print("Fetching WoodMac LNG export nominal capacity data...")
            capacity_df = fetch_capacity_data(conn)
            print(
                f"Retrieved {len(capacity_df):,} country-month rows from "
                f"{capacity_df['month'].min()} to {capacity_df['month'].max()}."
            )

            print(f"Building {CAPACITY_SHEET_NAME} matrix...")
            sheet_data[CAPACITY_SHEET_NAME] = build_country_matrix(capacity_df)

            print("Fetching WoodMac LNG export maintenance data...")
            maintenance_df = fetch_maintenance_data(conn)
            print(
                f"Retrieved {len(maintenance_df):,} country-month rows from "
                f"{maintenance_df['month'].min()} to {maintenance_df['month'].max()}."
            )

            print(f"Building {MAINTENANCE_SHEET_NAME} matrix...")
            sheet_data[MAINTENANCE_SHEET_NAME] = build_country_matrix(maintenance_df)
    finally:
        engine.dispose()

    print(f"Writing Excel file to {args.output}...")
    output_path = export_to_excel(sheet_data, args.output)

    print("Done.")
    print(f"Output file: {output_path}")

    for sheet_name, df in sheet_data.items():
        print(
            f"{sheet_name}: {len(df):,} rows | "
            f"Months: {df['Month'].min():%Y-%m} to {df['Month'].max():%Y-%m}"
        )


if __name__ == "__main__":
    main()
